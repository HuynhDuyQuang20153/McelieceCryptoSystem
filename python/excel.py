import openpyxl
import os


class Excel:
    def __init__(self, private_key, public_key, direc_cipherText, direc_plaintText, matrix):
        self.S = private_key['S']
        self.P = private_key['P']
        self.G = private_key['G']
        self.Gp = public_key['Gp']
        self.vector_cipher = direc_cipherText['vector']
        self.ciphertext = direc_cipherText['ciphertext']
        self.vector_plain = direc_plaintText['vector']
        self.plainText = direc_plaintText['plainText']
        self.matrix = matrix


    def Excel_Write(self):
        dir_path = os.path.join(os.getcwd(), "..", "data", "Excel_file")
        max_index = 0
        for filename in os.listdir(dir_path):
            if filename.startswith("Excel_folder_") and os.path.isdir(os.path.join(dir_path, filename)):
                index = int(filename.split("_")[2])
                if index > max_index:
                    max_index = index

        new_dir_path = os.path.join(dir_path, "Excel_folder_" + str(max_index + 1))
        os.mkdir(new_dir_path)

        workbook = openpyxl.Workbook()
        sheet_G = workbook.active
        sheet_G.title = "G"
        sheet_S = workbook.create_sheet("S")
        sheet_P = workbook.create_sheet("P")
        sheet_Gp = workbook.create_sheet("Gp")
        sheet_Cp = workbook.create_sheet("Cp")
        sheet_y = workbook.create_sheet("y")
        sheet_yy_ = workbook.create_sheet("yy_")
        sheet_x = workbook.create_sheet("x")
        sheet_matrix = workbook.create_sheet("matrix")
        file_name = "Excel.xlsx"
        data_path = os.path.join(new_dir_path, file_name)
        workbook.save(data_path)

        for i in range(self.G.shape[0]):
            for j in range(self.G.shape[1]):
                cell_G = sheet_G.cell(row=i+1, column=j+1)
                cell_G.value = self.G[i][j]
        print(f"\t- Save matrix G  in: {data_path}")

        for i in range(self.S.shape[0]):
            for j in range(self.S.shape[1]):
                cell_S = sheet_S.cell(row=i+1, column=j+1)
                cell_S.value = self.S[i][j]
        print(f"\t- Save matrix S  in: {data_path}")

        for i in range(self.P.shape[0]):
            for j in range(self.P.shape[1]):
                cell_P = sheet_P.cell(row=i+1, column=j+1)
                cell_P.value = self.P[i][j]
        print(f"\t- Save matrix P  in: {data_path}")

        for i in range(self.Gp.shape[0]):
            for j in range(self.Gp.shape[1]):
                cell_Gp = sheet_Gp.cell(row=i+1, column=j+1)
                cell_Gp.value = self.Gp[i][j]
        print(f"\t- Save matrix G' in: {data_path}")

        if self.vector_cipher.ndim == 1:
            for i in range(len(self.vector_cipher)):
                cell_c = sheet_Cp.cell(row=1, column=i+1)
                cell_c.value = self.vector_cipher[i]
            print(f"\t- Save vector C' in: {data_path}")

        elif self.vector_cipher.ndim > 1 or len(self.vector_cipher.shape) > 1:
            for i in range(self.vector_cipher.shape[0]):
                for j in range(self.vector_cipher.shape[1]):
                    cell_c = sheet_Cp.cell(row=i+1, column=j+1)
                    cell_c.value = self.vector_cipher[i][j]
            print(f"\t- Save vector C' in: {data_path}")

        for i in range(self.ciphertext.shape[0]):
            for j in range(self.ciphertext.shape[1]):
                cell_y = sheet_y.cell(row=i+1, column=j+1)
                cell_y.value = self.ciphertext[i][j]
        print(f"\t- Save cipher y  in: {data_path}")


        for i in range(self.vector_plain.shape[0]):
            for j in range(self.vector_plain.shape[1]):
                cell_yy_ = sheet_yy_.cell(row=i+1, column=j+1)
                cell_yy_.value = self.vector_plain[i][j]
        print(f"\t- Save vector y' in: {data_path}")

        if self.plainText.ndim == 1:
            for i in range(len(self.plainText)):
                cell_x = sheet_x.cell(row=1, column=i+1)
                cell_x.value = self.plainText[i]
            print(f"\t- Save plain x   in: {data_path}")

        elif self.plainText.ndim > 1 or len(self.plainText.shape) > 1:
            for i in range(self.plainText.shape[0]):
                for j in range(self.plainText.shape[1]):
                    cell_x = sheet_x.cell(row=i+1, column=j+1)
                    cell_x.value = self.plainText[i][j]
            print(f"\t- Save plain x   in: {data_path}")

        if self.matrix.ndim == 1:
            for i in range(len(self.matrix)):
                cell_matrix = sheet_matrix.cell(row=1, column=i+1)
                cell_matrix.value = self.matrix[i]
            print(f"\t- Save matrix    in {data_path}")

        elif self.matrix.ndim > 1 or len(self.matrix.shape) > 1:
            for i in range(self.matrix.shape[0]):
                for j in range(self.matrix.shape[1]):
                    cell_matrix = sheet_matrix.cell(row=i+1, column=j+1)
                    cell_matrix.value = self.matrix[i][j]
            print(f"\t- Save matrix    in: {data_path}")

        workbook.save(data_path)
